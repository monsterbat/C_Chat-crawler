"""
Crawler C_Chat
Latest:2024/05/14
"""
# 
from basic_function import getData as getData
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

def c_chat_z(keyword, date, search_type):
    pageurl = "https://www.ptt.cc/bbs/C_Chat/index.html"
    keyword_ori = keyword
    keyword = keyword.lower()
    keyword = keyword.split()
    keyword_len = []
    for keyword_ls in keyword:
        keyword_len = keyword_len + [len(keyword_ls)]

    keyword_data = {"keyword_data":[]}
    for keyword_ls,keyword_len_ls in zip(keyword,keyword_len):
        keyword_data["keyword_data"].append({"keyword":keyword_ls,"keyword_len":keyword_len_ls})

    # Choose data need analysis
    # 輸入日期
    date_year = int(date.split("-")[0])
    date_month = int(date.split("-")[1])
    date_day = int(date.split("-")[2])
    # 當下日期
    now = datetime.now()
    now_year = int(now.year)
    now_month = int(now.month)
    now_day = int(now.day)

    switch = True
    # Creat file and clean previous data
    output_json_data = {
        "result":[
            {"title":"標題","date":"日期"}
        ]
    }

    while switch == True:
        # 
        data_result = getData(pageurl)
        # 取出日期驗證
        page_date = data_result["result"]["require_data"][0]["date"]
        page_year = int(now_year) # 尚未進行跨年處理！
        page_month = int(page_date.split("/")[0])
        page_day = int(page_date.split("/")[1])

        print("---1---",switch)
        if date_year > page_year:
            switch = False
        if date_year == page_year:
            if date_month > page_month:
                switch = False
            if date_month == page_month:
                if date_day > page_day:
                    switch = False

        if switch == True:            
            require_data_result = data_result["result"]["require_data"]
            output_data = analysis_data(require_data_result,search_type,keyword_data)
            for output_data_ls in output_data:
                output_json_data["result"].append(output_data_ls)

        pageurl="https://www.ptt.cc" + data_result["result"]["url"]#換頁
    if switch == False:
        # 寫入xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = keyword_ori
        xlsx_data = output_json_data["result"]
        print(xlsx_data)
        for index, entry in enumerate(xlsx_data, start=1):
            ws[f'A{index}'] = entry["title"]
            ws[f'B{index}'] = entry["date"]
        # 建立總表
        wb = load_workbook("data.xlsx")
        summary_sheet = wb.create_sheet(title="Summary")
        summary_data_json = {
            "result":[
                {"keyword":"keyword","count":"討論數"},
                {"keyword":keyword_ori,"count":len(xlsx_data)}
            ]
        }
        summary_data = summary_data_json["result"]
        for index, entry in enumerate(summary_data, start=1):
            summary_sheet[f'A{index}'] = entry["keyword"]
            summary_sheet[f'B{index}'] = entry["count"]
        wb.save("data.xlsx")
        print("Finish!")
    

def analysis_data(require_data_result,search_type,keyword_data):    
    
    # 取得資料變小寫
    output_data = []
    require_data_result_lower = require_data_result
    for title_result_ls in require_data_result_lower:
        title_result_ls["title"] = title_result_ls['title'].lower()

    # 找標題
    if search_type == "title only":
        for require_data_result_lower_ls in require_data_result_lower:
            # print("!!!!!!!!",type(movie_title_filter))
            title_result_lower = require_data_result_lower_ls["title"]

            for keyword_data_ls in keyword_data["keyword_data"]:
                keyword = keyword_data_ls["keyword"]
                keyword_len = keyword_data_ls["keyword_len"]
                if title_result_lower[1:keyword_len+1] == keyword:
                    output_data = output_data + [require_data_result_lower_ls]

    # 模糊搜尋         
    if search_type == "include all":  
        for require_data_result_lower_ls in require_data_result_lower:
            title_result_lower = require_data_result_lower_ls["title"]

            for keyword_data_ls in keyword_data["keyword_data"]:
                keyword = keyword_data_ls["keyword"]
                keyword_len = keyword_data_ls["keyword_len"]

                if keyword in title_result_lower:
                    output_data = output_data + [require_data_result_lower_ls]

    print("***",output_data)
    return output_data
